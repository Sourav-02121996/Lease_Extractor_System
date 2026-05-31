"""Backend test suite for the Lease Abstraction Assistant.

Covers:
- /api/health
- /api/documents (list, stats, get-by-id)
- /api/upload (real PDF via Gemini extraction)
- /api/documents/{id}/draft and /approve
- /api/export/approved (CSV + empty-state message)
"""
import io
import os
import time

import pytest
import requests
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen import canvas

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    # Fall back to the value already baked into the frontend env for local runs
    with open("/app/frontend/.env") as fh:
        for line in fh:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip().rstrip("/")
                break

assert BASE_URL, "REACT_APP_BACKEND_URL must be set"

API = f"{BASE_URL}/api"

FIELD_NAMES = [
    "Landlord / Property Owner Name",
    "Tenant / Business Name",
    "Guarantor Name(s)",
    "Mailing Addresses for all parties",
    "Contact Information",
    "Effective Date / Lease Start Date",
    "Lease End Date",
    "Lease Length",
    "Renewal Option Details",
    "Holdover Terms",
]


@pytest.fixture(scope="session")
def session():
    s = requests.Session()
    s.headers.update({"Accept": "application/json"})
    return s


# ---------- Health ----------
class TestHealth:
    def test_health_ok(self, session):
        r = session.get(f"{API}/health", timeout=15)
        assert r.status_code == 200
        data = r.json()
        assert data["status"] == "ok"
        assert data["ai_key_configured"] is True


# ---------- Documents listing / stats / detail ----------
class TestDocuments:
    def test_list_returns_seed_docs(self, session):
        r = session.get(f"{API}/documents", timeout=15)
        assert r.status_code == 200
        docs = r.json()
        assert isinstance(docs, list)
        assert len(docs) >= 6, f"expected at least 6 seeded docs, got {len(docs)}"
        for d in docs[:3]:
            assert "id" in d and "file_name" in d and "status" in d
            assert isinstance(d.get("fields"), list)

    def test_stats_shape_and_counts(self, session):
        r = session.get(f"{API}/documents/stats", timeout=15)
        assert r.status_code == 200
        data = r.json()
        for k in ["total", "processed", "needs_review", "approved"]:
            assert k in data and isinstance(data[k], int)
        assert data["total"] >= 6
        assert data["approved"] >= 1

    def test_get_single_doc_has_10_fields_and_raw_text(self, session):
        listing = session.get(f"{API}/documents", timeout=15).json()
        doc_id = listing[0]["id"]
        r = session.get(f"{API}/documents/{doc_id}", timeout=15)
        assert r.status_code == 200
        doc = r.json()
        assert "raw_text" in doc
        assert len(doc["fields"]) == 10
        names = {f["fieldName"] for f in doc["fields"]}
        assert names == set(FIELD_NAMES)
        for f in doc["fields"]:
            for key in ("fieldName", "value", "confidence", "evidence", "status", "section"):
                assert key in f

    def test_get_invalid_id_returns_400(self, session):
        r = session.get(f"{API}/documents/not-an-oid", timeout=15)
        assert r.status_code == 400

    def test_get_unknown_id_returns_404(self, session):
        r = session.get(f"{API}/documents/507f1f77bcf86cd799439011", timeout=15)
        assert r.status_code == 404


# ---------- Upload + AI extraction ----------
LEASE_TEXT_LINES = [
    "COMMERCIAL LEASE AGREEMENT",
    "",
    "This Lease Agreement is made effective as of March 1, 2024 between:",
    "Landlord: Acme Real Estate Holdings LLC, a Delaware limited liability company,",
    "with offices at 123 Market Street, Suite 900, San Francisco, CA 94105.",
    "Phone: (415) 555-0123. Email: leasing@acmere.com.",
    "",
    "Tenant: Skyline Robotics Inc., a California corporation,",
    "with offices at 500 Innovation Way, Palo Alto, CA 94301.",
    "",
    "Guarantor: John A. Sterling, an individual residing in San Francisco, CA.",
    "",
    "TERM. The initial term of this Lease shall be five (5) years, commencing on",
    "March 1, 2024 (the 'Commencement Date') and expiring on February 28, 2029.",
    "",
    "RENEWAL. Tenant shall have one (1) option to renew for an additional",
    "five (5) year term at then-current fair market rent, with written notice",
    "given not less than 180 days prior to expiration.",
    "",
    "HOLDOVER. If Tenant remains in possession after expiration, such tenancy",
    "shall be month-to-month at 150% of the then-current Base Rent.",
    "",
    "NOTICES. All notices shall be sent to the addresses set forth above.",
]


def _make_lease_pdf_bytes() -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=LETTER)
    text = c.beginText(72, 720)
    text.setFont("Helvetica", 11)
    for line in LEASE_TEXT_LINES:
        text.textLine(line)
    c.drawText(text)
    c.showPage()
    c.save()
    return buf.getvalue()


@pytest.fixture(scope="session")
def uploaded_doc(session):
    pdf_bytes = _make_lease_pdf_bytes()
    files = {"file": ("TEST_lease.pdf", pdf_bytes, "application/pdf")}
    r = session.post(f"{API}/upload", files=files, timeout=120)
    assert r.status_code == 200, f"upload failed: {r.status_code} {r.text[:400]}"
    return r.json()


class TestUpload:
    def test_upload_rejects_non_pdf(self, session):
        files = {"file": ("foo.txt", b"hello world", "text/plain")}
        r = session.post(f"{API}/upload", files=files, timeout=30)
        assert r.status_code == 400

    def test_upload_text_pdf_extracts_all_fields(self, uploaded_doc):
        doc = uploaded_doc
        assert doc["extraction_method"] == "text_pdf"
        assert doc["status"] in ("processed", "needs_review")
        assert len(doc["fields"]) == 10
        names = {f["fieldName"] for f in doc["fields"]}
        assert names == set(FIELD_NAMES)
        # at least some core fields extracted with value present
        by_name = {f["fieldName"]: f for f in doc["fields"]}
        for required in [
            "Landlord / Property Owner Name",
            "Tenant / Business Name",
            "Effective Date / Lease Start Date",
            "Lease End Date",
        ]:
            assert by_name[required]["value"], f"expected {required} extracted"
            assert by_name[required]["status"] in ("extracted", "needs_review", "approved")
            assert 0.0 <= by_name[required]["confidence"] <= 1.0
            assert by_name[required]["evidence"]

    def test_upload_does_not_invent_missing(self, uploaded_doc):
        # Our test PDF does NOT contain a phone+email block tied to Tenant, but
        # the more important invariant: any field returned with empty value must
        # be marked missing with confidence 0.
        for f in uploaded_doc["fields"]:
            if not f["value"]:
                assert f["status"] == "missing"
                assert f["confidence"] in (0, 0.0)


# ---------- Draft + Approve ----------
class TestDraftAndApprove:
    def test_save_draft_updates_value(self, session, uploaded_doc):
        doc_id = uploaded_doc["id"]
        payload = {
            "fields": [
                {
                    "fieldName": "Holdover Terms",
                    "value": "TEST_DRAFT 150% holdover month-to-month",
                    "status": "needs_review",
                }
            ]
        }
        r = session.put(f"{API}/documents/{doc_id}/draft", json=payload, timeout=30)
        assert r.status_code == 200, r.text
        updated = r.json()
        f = next(x for x in updated["fields"] if x["fieldName"] == "Holdover Terms")
        assert f["value"] == "TEST_DRAFT 150% holdover month-to-month"
        # Re-fetch to confirm persistence
        again = session.get(f"{API}/documents/{doc_id}", timeout=15).json()
        f2 = next(x for x in again["fields"] if x["fieldName"] == "Holdover Terms")
        assert f2["value"] == "TEST_DRAFT 150% holdover month-to-month"

    def test_approve_promotes_status(self, session, uploaded_doc):
        doc_id = uploaded_doc["id"]
        r = session.post(f"{API}/documents/{doc_id}/approve", timeout=30)
        assert r.status_code == 200, r.text
        updated = r.json()
        assert updated["status"] == "approved"
        # all non-empty fields should be marked approved
        for f in updated["fields"]:
            if f["value"]:
                assert f["status"] == "approved", f"{f['fieldName']} not approved"

    def test_approved_doc_cannot_be_drafted(self, session, uploaded_doc):
        doc_id = uploaded_doc["id"]
        r = session.put(
            f"{API}/documents/{doc_id}/draft",
            json={"fields": [{"fieldName": "Holdover Terms", "value": "x"}]},
            timeout=15,
        )
        assert r.status_code == 400


# ---------- Export ----------
class TestExport:
    def test_export_returns_csv_with_14_columns(self, session):
        r = session.get(f"{API}/export/approved", timeout=30)
        assert r.status_code == 200, r.text
        assert "text/csv" in r.headers.get("content-type", "")
        body = r.text
        lines = [ln for ln in body.splitlines() if ln.strip()]
        assert len(lines) >= 2  # header + at least one approved row
        import csv as _csv
        reader = _csv.reader(io.StringIO(body))
        rows = list(reader)
        assert len(rows[0]) == 14, f"expected 14 columns, got {len(rows[0])}: {rows[0]}"
        # data row column count should match
        assert len(rows[1]) == 14

    def test_approved_count_endpoint(self, session):
        r = session.get(f"{API}/export/approved/count", timeout=15)
        assert r.status_code == 200
        data = r.json()
        assert "approved" in data and isinstance(data["approved"], int)
        assert data["approved"] >= 1

    def test_excel_export_returns_workbook_with_14_columns(self, session):
        r = session.get(f"{API}/export/approved/excel", timeout=30)
        assert r.status_code == 200, r.text
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in r.headers.get(
            "content-type", ""
        )
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(r.content), read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        assert len(rows) >= 2  # header + at least one approved row
        assert len(rows[0]) == 14
        assert len(rows[1]) == 14

    def test_empty_export_message_when_no_approved(self, session):
        # Mark all approved docs as needs_review temporarily, then restore.
        listing = session.get(f"{API}/documents", timeout=15).json()
        approved_ids = [d["id"] for d in listing if d["status"] == "approved"]
        assert approved_ids, "need at least one approved doc to run this test"

        # Use direct mongo via a tiny admin trick: we don't have such endpoint,
        # so simulate by toggling status via approve route is not enough.
        # Instead, connect to mongo directly.
        from pymongo import MongoClient
        from bson import ObjectId

        mongo_url = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
        db_name = os.environ.get("DB_NAME", "lease_abstraction")
        client = MongoClient(mongo_url)
        col = client[db_name]["documents"]
        try:
            col.update_many(
                {"_id": {"$in": [ObjectId(i) for i in approved_ids]}},
                {"$set": {"status": "needs_review"}},
            )
            r = session.get(f"{API}/export/approved", timeout=15)
            assert r.status_code == 404
            assert "No approved lease records available for export." in r.text
        finally:
            col.update_many(
                {"_id": {"$in": [ObjectId(i) for i in approved_ids]}},
                {"$set": {"status": "approved"}},
            )
            client.close()


# ---------- Cleanup uploaded test doc ----------
@pytest.fixture(scope="session", autouse=True)
def _cleanup(request, session):
    yield
    # Remove TEST_ uploaded docs to keep DB tidy
    try:
        from pymongo import MongoClient
        mongo_url = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
        db_name = os.environ.get("DB_NAME", "lease_abstraction")
        client = MongoClient(mongo_url)
        client[db_name]["documents"].delete_many({"file_name": {"$regex": "^TEST_"}})
        client.close()
    except Exception as e:  # noqa
        print("cleanup failed:", e)
